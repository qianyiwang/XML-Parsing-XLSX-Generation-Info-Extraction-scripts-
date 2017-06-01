# data process step 3 is used to extract features from each subject form

import pandas as pd
import math
from openpyxl import load_workbook
import os

# define functions
def calculateAbsTime(time):
    if ':' not in time:
        return 0
    arr = time.split(':')
    absTime =  (int(arr[0])-10)*3600+int(arr[1])*60+int(arr[2])
    return absTime

def findSteps(startTime, endTime):
    steps = -1
    if startTime == endTime:
        return 1
    for t in df_event['SimulatorTime']:
        t = calculateAbsTime(t)
        if t >= startTime and t <= endTime:
            steps = steps + 1

    return steps

def findBaselineSteps(startTime, endTime):
    steps = 1
    for t in df_baselineEvent['SimulatorTime']:
        t = calculateAbsTime(t)
        if t >= startTime and t <= endTime:
            steps = steps + 1

    return steps

def calculateMean(startTime, endTime, feature):
    sumup = 0
    number = 0
    startFlag = False

    for i in range(0, len(df_data['SimulatorTime'])):
        if ':' in str(df_data['SimulatorTime'][i]):
            t = calculateAbsTime(str(df_data['SimulatorTime'][i]))
            if t>= startTime and startFlag == False:
                startVal = df_data[feature][i]
            if t >= startTime and t <= endTime:
                number += 1
                sumup += df_data[feature][i]

    if number == 0:
        if startTime == endTime:
            return startVal
        # else:
        #     return 1
    return sumup / number

def calculateBaselineMean(startTime, endTime, feature):
    sumup = 0
    number = 0
    for i in range(0, len(df_baselineData['SimulatorTime'])):
        if ':' in str(df_baselineData['SimulatorTime'][i]):
            t = calculateAbsTime(str(df_baselineData['SimulatorTime'][i]))
            if t >= startTime and t <= endTime:
                number += 1
                sumup += df_baselineData[feature][i]
    if number == 0:
        return 1

    return sumup / number

def findADASNum(startTime, endTime):
    number = 0
    for i in range(0, len(df_data['SimulatorTime'])):
        if ':' in str(df_data['SimulatorTime'][i]):
            t = calculateAbsTime(str(df_data['SimulatorTime'][i]))
            if t >= startTime and t <= endTime and 'ADAS' in str(df_data['ADAS'][i]):
                # print df_data['ADAS'][i]
                number += 1

    return number

def findBaselineADASNum(startTime, endTime):
    number = 0
    for i in range(0, len(df_baselineData['SimulatorTime'])):
        if ':' in str(df_baselineData['SimulatorTime'][i]):
            t = calculateAbsTime(str(df_baselineData['SimulatorTime'][i]))
            if t >= startTime and t <= endTime and 'ADAS' in str(df_baselineData['ADAS'][i]):
                number += 1

    return number


FILE = 'C:\ClinicServer\Excel'
for subdir, dirs, files in os.walk(FILE):
    for f in files:
        if 'xlsx' not in f:
            continue
        print 'processing on %s' % f
        fileDir = 'C:\ClinicServer\Excel\%s'%f
        book = pd.ExcelFile(fileDir)

        # 'Data' and 'Baseline Data' sheets processing to find event start time
        df_data = book.parse('Data')
        df_baselineData = book.parse('BaselineData')
        road_id_init = '00'
        road_id_array = []
        for i in range(0, len(df_data['RoadID'])):
            road_id = df_data['RoadID'][i]
            if road_id != road_id_init and not math.isnan(road_id):
                road_id_array.append(i)
                road_id_init = road_id
        end2end_mcs_startTime = df_data['SimulatorTime'][road_id_array[1]]
        end2end_mcs_startTime = calculateAbsTime(end2end_mcs_startTime)
        end2end_climate_startTime = df_data['SimulatorTime'][road_id_array[3]]
        end2end_climate_startTime = calculateAbsTime(end2end_climate_startTime)
        end2end_bt_startTime = df_data['SimulatorTime'][road_id_array[4]]
        end2end_bt_startTime = calculateAbsTime(end2end_bt_startTime)
        # end2end_park_startTime = df_data['SimulatorTime'][road_id_array[9]]

        road_id_init = '00'
        road_id_array = []
        for i in range(0, len(df_baselineData['RoadID'])):
            road_id = df_baselineData['RoadID'][i]
            if road_id != road_id_init and not math.isnan(road_id):
                road_id_array.append(i)
                road_id_init = road_id
        baseline_mcs_startTime = df_baselineData['SimulatorTime'][road_id_array[1]]
        baseline_mcs_startTime = calculateAbsTime(baseline_mcs_startTime)
        baseline_climate_startTime = df_baselineData['SimulatorTime'][road_id_array[3]]
        baseline_climate_startTime = calculateAbsTime(baseline_climate_startTime)
        baseline_bt_startTime = df_baselineData['SimulatorTime'][road_id_array[4]]
        baseline_bt_startTime = calculateAbsTime(baseline_bt_startTime)
        # baseline_park_startTime = df_baselineData['SimulatorTime'][road_id_array[9]]

        #  'Event' and 'Baseline Event' sheets processing to find event end time
        df_event = book.parse('Event')
        df_baselineEvent = book.parse('BaselineEvent')
        end2end_map_startTime = 'Nan'
        park_complete_flag = False
        map_complete_flag = False
        for i in range(0, len(df_event['Name'])):
            if df_event['Name'][i] == 'Start Google Map' and map_complete_flag == False:
                map_complete_flag = True
                end2end_map_endTime = df_event['SimulatorTime'][i]
                end2end_map_endTime = calculateAbsTime(end2end_map_endTime)
            if df_event['Name'][i] == 'EnterAddressIssued':
                end2end_map_startTime = df_event['SimulatorTime'][i]
                end2end_map_startTime = calculateAbsTime(end2end_map_startTime)
            if df_event['Name'][i] == 'driver_tempUp':
                end2end_climate_endTime = df_event['SimulatorTime'][i]
                end2end_climate_endTime = calculateAbsTime(end2end_climate_endTime)
            if df_event['Name'][i] == 'audio_seekUp':
                end2end_bt_endTime = df_event['SimulatorTime'][i]
                end2end_bt_endTime = calculateAbsTime(end2end_bt_endTime)
            if df_event['Name'][i] == 'MCS_enter':
                end2end_mcs_endTime = df_event['SimulatorTime'][i]
                end2end_mcs_endTime = calculateAbsTime(end2end_mcs_endTime)
            if df_event['Name'][i] == 'Parking Notification':
                end2end_park_startTime = df_event['SimulatorTime'][i]
                end2end_park_startTime = calculateAbsTime(end2end_park_startTime)
            if df_event['Name'][i] == 'PARKING' and park_complete_flag == False:
                park_complete_flag = True
                end2end_park_endTime = df_event['SimulatorTime'][i]
                end2end_park_endTime = calculateAbsTime(end2end_park_endTime)
        if end2end_map_startTime == 'Nan':
            end2end_map_startTime = end2end_map_endTime

        baseline_map_startTime = 'Nan'
        park_complete_flag = False
        for i in range(0, len(df_baselineEvent['Name'])):
            if df_baselineEvent['Name'][i] == 'EnterAddressExecuted':
                baseline_map_endTime = df_baselineEvent['SimulatorTime'][i]
                baseline_map_endTime = calculateAbsTime(baseline_map_endTime)
            if df_baselineEvent['Name'][i] == 'EnterAddressIssued':
                baseline_map_startTime = df_baselineEvent['SimulatorTime'][i]
                baseline_map_startTime = calculateAbsTime(baseline_map_startTime)
            if df_baselineEvent['Name'][i] == 'climate_dvr_tempUp':
                baseline_climate_endTime = df_baselineEvent['SimulatorTime'][i]
                baseline_climate_endTime = calculateAbsTime(baseline_climate_endTime)
            if df_baselineEvent['Name'][i] == 'audio_seekup':
                baseline_bt_endTime = df_baselineEvent['SimulatorTime'][i]
                baseline_bt_endTime = calculateAbsTime(baseline_bt_endTime)
            if df_baselineEvent['Name'][i] == 'MCS clicked':
                baseline_mcs_endTime = df_baselineEvent['SimulatorTime'][i]
                baseline_mcs_endTime = calculateAbsTime(baseline_mcs_endTime)
            if df_baselineEvent['Name'][i] == 'parkopediaIssued':
                baseline_park_startTime = df_baselineEvent['SimulatorTime'][i]
                baseline_park_startTime = calculateAbsTime(baseline_park_startTime)
            if df_baselineEvent['Name'][i] == 'parkopediaExecuted' and park_complete_flag == False:
                park_complete_flag = True
                baseline_park_endTime = df_baselineEvent['SimulatorTime'][i]
                baseline_park_endTime = calculateAbsTime(baseline_park_endTime)
        if baseline_map_startTime == 'Nan':
            baseline_map_startTime = baseline_map_endTime

        # find task steps
        end2end_map_step = findSteps(end2end_map_startTime, end2end_map_endTime)
        end2end_mcs_step = findSteps(end2end_mcs_startTime, end2end_mcs_endTime)
        end2end_bt_step = findSteps(end2end_bt_startTime, end2end_bt_endTime)
        end2end_climate_step = findSteps(end2end_climate_startTime, end2end_climate_endTime)
        end2end_park_step = findSteps(end2end_park_startTime, end2end_park_endTime)

        baseline_map_step = findBaselineSteps(baseline_map_startTime, baseline_map_endTime)
        baseline_mcs_step = findBaselineSteps(baseline_mcs_startTime, baseline_mcs_endTime)
        baseline_bt_step = findBaselineSteps(baseline_bt_startTime, baseline_bt_endTime)
        baseline_climate_step = findBaselineSteps(baseline_climate_startTime, baseline_climate_endTime)
        baseline_park_step = findBaselineSteps(baseline_park_startTime, baseline_park_endTime)

        # calculate velocity mean
        end2end_map_velocity_mean = calculateMean(end2end_map_startTime, end2end_map_endTime, 'Velocity')
        end2end_mcs_velocity_mean = calculateMean(end2end_mcs_startTime, end2end_mcs_endTime, 'Velocity')
        end2end_climate_velocity_mean = calculateMean(end2end_climate_startTime, end2end_climate_endTime, 'Velocity')
        end2end_bt_velocity_mean = calculateMean(end2end_bt_startTime, end2end_bt_endTime, 'Velocity')
        end2end_park_velocity_mean = calculateMean(end2end_park_startTime, end2end_park_endTime, 'Velocity')

        baseline_map_velocity_mean = calculateBaselineMean(baseline_map_startTime, baseline_map_endTime, 'Velocity')
        baseline_mcs_velocity_mean = calculateBaselineMean(baseline_mcs_startTime, baseline_mcs_endTime, 'Velocity')
        baseline_climate_velocity_mean = calculateBaselineMean(baseline_climate_startTime, baseline_climate_endTime, 'Velocity')
        baseline_bt_velocity_mean = calculateBaselineMean(baseline_bt_startTime, baseline_bt_endTime, 'Velocity')
        baseline_park_velocity_mean = calculateBaselineMean(baseline_park_startTime, baseline_park_endTime, 'Velocity')

        # calculate acceleration mean
        end2end_map_acc_mean = calculateMean(end2end_map_startTime, end2end_map_endTime, 'Acceleration')
        end2end_mcs_acc_mean = calculateMean(end2end_mcs_startTime, end2end_mcs_endTime, 'Acceleration')
        end2end_climate_acc_mean = calculateMean(end2end_climate_startTime, end2end_climate_endTime, 'Acceleration')
        end2end_bt_acc_mean = calculateMean(end2end_bt_startTime, end2end_bt_endTime, 'Acceleration')
        end2end_park_acc_mean = calculateMean(end2end_park_startTime, end2end_park_endTime, 'Acceleration')

        baseline_map_acc_mean = calculateBaselineMean(baseline_map_startTime, baseline_map_endTime, 'Acceleration')
        baseline_mcs_acc_mean = calculateBaselineMean(baseline_mcs_startTime, baseline_mcs_endTime, 'Acceleration')
        baseline_climate_acc_mean = calculateBaselineMean(baseline_climate_startTime, baseline_climate_endTime, 'Acceleration')
        baseline_bt_acc_mean = calculateBaselineMean(baseline_bt_startTime, baseline_bt_endTime, 'Acceleration')
        baseline_park_acc_mean = calculateBaselineMean(baseline_park_startTime, baseline_park_endTime, 'Acceleration')

        # calculate brake state mean
        end2end_map_bs_mean = calculateMean(end2end_map_startTime, end2end_map_endTime, 'BrakeState')
        end2end_mcs_bs_mean = calculateMean(end2end_mcs_startTime, end2end_mcs_endTime, 'BrakeState')
        end2end_climate_bs_mean = calculateMean(end2end_climate_startTime, end2end_climate_endTime, 'BrakeState')
        end2end_bt_bs_mean = calculateMean(end2end_bt_startTime, end2end_bt_endTime, 'BrakeState')
        end2end_park_bs_mean = calculateMean(end2end_park_startTime, end2end_park_endTime, 'BrakeState')

        baseline_map_bs_mean = calculateBaselineMean(baseline_map_startTime, baseline_map_endTime, 'BrakeState')
        baseline_mcs_bs_mean = calculateBaselineMean(baseline_mcs_startTime, baseline_mcs_endTime, 'BrakeState')
        baseline_climate_bs_mean = calculateBaselineMean(baseline_climate_startTime, baseline_climate_endTime, 'BrakeState')
        baseline_bt_bs_mean = calculateBaselineMean(baseline_bt_startTime, baseline_bt_endTime, 'BrakeState')
        baseline_park_bs_mean = calculateBaselineMean(baseline_park_startTime, baseline_park_endTime, 'BrakeState')

        # calculate steering wheel angle mean
        end2end_map_swa_mean = calculateMean(end2end_map_startTime, end2end_map_endTime, 'SteeringWheelAngle')
        end2end_mcs_swa_mean = calculateMean(end2end_mcs_startTime, end2end_mcs_endTime, 'SteeringWheelAngle')
        end2end_climate_swa_mean = calculateMean(end2end_climate_startTime, end2end_climate_endTime, 'SteeringWheelAngle')
        end2end_bt_swa_mean = calculateMean(end2end_bt_startTime, end2end_bt_endTime, 'SteeringWheelAngle')
        end2end_park_swa_mean = calculateMean(end2end_park_startTime, end2end_park_endTime, 'SteeringWheelAngle')

        baseline_map_swa_mean = calculateBaselineMean(baseline_map_startTime, baseline_map_endTime, 'SteeringWheelAngle')
        baseline_mcs_swa_mean = calculateBaselineMean(baseline_mcs_startTime, baseline_mcs_endTime, 'SteeringWheelAngle')
        baseline_climate_swa_mean = calculateBaselineMean(baseline_climate_startTime, baseline_climate_endTime, 'SteeringWheelAngle')
        baseline_bt_swa_mean = calculateBaselineMean(baseline_bt_startTime, baseline_bt_endTime, 'SteeringWheelAngle')
        baseline_park_swa_mean = calculateBaselineMean(baseline_park_startTime, baseline_park_endTime, 'SteeringWheelAngle')

        # calculate lane offset mean
        end2end_map_laneoff_mean = calculateMean(end2end_map_startTime, end2end_map_endTime, 'LaneOffset')
        end2end_mcs_laneoff_mean = calculateMean(end2end_mcs_startTime, end2end_mcs_endTime, 'LaneOffset')
        end2end_climate_laneoff_mean = calculateMean(end2end_climate_startTime, end2end_climate_endTime, 'LaneOffset')
        end2end_bt_laneoff_mean = calculateMean(end2end_bt_startTime, end2end_bt_endTime, 'LaneOffset')
        end2end_park_laneoff_mean = calculateMean(end2end_park_startTime, end2end_park_endTime, 'LaneOffset')

        baseline_map_laneoff_mean = calculateBaselineMean(baseline_map_startTime, baseline_map_endTime, 'LaneOffset')
        baseline_mcs_laneoff_mean = calculateBaselineMean(baseline_mcs_startTime, baseline_mcs_endTime, 'LaneOffset')
        baseline_climate_laneoff_mean = calculateBaselineMean(baseline_climate_startTime, baseline_climate_endTime, 'LaneOffset')
        baseline_bt_laneoff_mean = calculateBaselineMean(baseline_bt_startTime, baseline_bt_endTime, 'LaneOffset')
        baseline_park_laneoff_mean = calculateBaselineMean(baseline_park_startTime, baseline_park_endTime, 'LaneOffset')

        # calculate HR mean
        end2end_map_hr_mean = calculateMean(end2end_map_startTime, end2end_map_endTime, 'HR')
        end2end_mcs_hr_mean = calculateMean(end2end_mcs_startTime, end2end_mcs_endTime, 'HR')
        end2end_climate_hr_mean = calculateMean(end2end_climate_startTime, end2end_climate_endTime, 'HR')
        end2end_bt_hr_mean = calculateMean(end2end_bt_startTime, end2end_bt_endTime, 'HR')
        end2end_park_hr_mean = calculateMean(end2end_park_startTime, end2end_park_endTime, 'HR')

        baseline_map_hr_mean = calculateBaselineMean(baseline_map_startTime, baseline_map_endTime, 'HR')
        baseline_mcs_hr_mean = calculateBaselineMean(baseline_mcs_startTime, baseline_mcs_endTime, 'HR')
        baseline_climate_hr_mean = calculateBaselineMean(baseline_climate_startTime, baseline_climate_endTime, 'HR')
        baseline_bt_hr_mean = calculateBaselineMean(baseline_bt_startTime, baseline_bt_endTime, 'HR')
        baseline_park_hr_mean = calculateBaselineMean(baseline_park_startTime, baseline_park_endTime, 'HR')

        # calculate ADAS number
        end2end_map_adasNum = findADASNum(end2end_map_startTime, end2end_map_endTime)
        end2end_mcs_adasNum = findADASNum(end2end_mcs_startTime, end2end_mcs_endTime)
        end2end_climate_adasNum = findADASNum(end2end_climate_startTime, end2end_climate_endTime)
        end2end_bt_adasNum = findADASNum(end2end_bt_startTime, end2end_bt_endTime)
        end2end_park_adasNum = findADASNum(end2end_park_startTime, end2end_park_endTime)

        baseline_map_adasNum = findBaselineADASNum(baseline_map_startTime, baseline_map_endTime)
        baseline_mcs_adasNum = findBaselineADASNum(baseline_mcs_startTime, baseline_mcs_endTime)
        baseline_climate_adasNum = findBaselineADASNum(baseline_climate_startTime, baseline_climate_endTime)
        baseline_bt_adasNum = findBaselineADASNum(baseline_bt_startTime, baseline_bt_endTime)
        baseline_park_adasNum = findBaselineADASNum(baseline_park_startTime, baseline_park_endTime)



        # write into 'Data' sheet
        book = load_workbook(fileDir)
        writer = pd.ExcelWriter(fileDir, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        update = pd.DataFrame(
        {
            'Task Name':['Open map', 'MCS', 'Climate control', 'BT audio control', 'Parking'],
            'Task Start Time':[end2end_map_startTime, end2end_mcs_startTime, end2end_climate_startTime, end2end_bt_startTime, end2end_park_startTime],
            'Task End Time':[end2end_map_endTime, end2end_mcs_endTime, end2end_climate_endTime, end2end_bt_endTime, end2end_park_endTime],
            'Task Duration': [end2end_map_endTime-end2end_map_startTime, end2end_mcs_endTime-end2end_mcs_startTime, end2end_climate_endTime-end2end_climate_startTime, end2end_bt_endTime-end2end_bt_startTime, end2end_park_endTime-end2end_park_startTime],
            'Task Steps': [end2end_map_step, end2end_mcs_step, end2end_climate_step, end2end_bt_step, end2end_park_step],
            'Average HR': [end2end_map_hr_mean, end2end_mcs_hr_mean, end2end_climate_hr_mean, end2end_bt_hr_mean, end2end_park_hr_mean],
            'Average Velocity': [end2end_map_velocity_mean, end2end_mcs_velocity_mean, end2end_climate_velocity_mean, end2end_bt_velocity_mean, end2end_park_velocity_mean],
            'Average Acceleration': [end2end_map_acc_mean, end2end_mcs_acc_mean, end2end_climate_acc_mean, end2end_bt_acc_mean, end2end_park_acc_mean],
            'Average BrakeStatus': [end2end_map_bs_mean, end2end_mcs_bs_mean, end2end_climate_bs_mean, end2end_bt_bs_mean, end2end_park_bs_mean],
            'Average LaneOffset': [end2end_map_laneoff_mean, end2end_mcs_laneoff_mean, end2end_climate_laneoff_mean, end2end_bt_laneoff_mean, end2end_park_laneoff_mean],
            'Average SteeringWheelAngle': [end2end_map_swa_mean, end2end_mcs_swa_mean, end2end_climate_swa_mean, end2end_bt_swa_mean, end2end_park_swa_mean],
            'ADAS#': [end2end_map_adasNum, end2end_mcs_adasNum, end2end_climate_adasNum, end2end_bt_adasNum, end2end_park_adasNum]
        })
        update.to_excel(writer, 'Task',
            columns=['Task Name','Task Start Time', 'Task End Time', 'Task Duration', 'Task Steps', 'Average HR', 'Average Velocity', 'Average Acceleration', 'Average BrakeStatus', 'Average LaneOffset', 'Average SteeringWheelAngle', 'ADAS#'],
            index=False)
        writer.save()


        # write into 'Baseline Data' sheet
        book = load_workbook(fileDir)
        writer = pd.ExcelWriter(fileDir, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        update = pd.DataFrame(
        {
            'Task Name':['Open map', 'MCS', 'Climate control', 'BT audio control', 'Parking'],
            'Task Start Time':[baseline_map_startTime, baseline_mcs_startTime, baseline_climate_startTime, baseline_bt_startTime, baseline_park_startTime],
            'Task End Time':[baseline_map_endTime, baseline_mcs_endTime, baseline_climate_endTime, baseline_bt_endTime, baseline_park_endTime],
            'Task Duration': [baseline_map_endTime-baseline_map_startTime, baseline_mcs_endTime-baseline_mcs_startTime, baseline_climate_endTime-baseline_climate_startTime, baseline_bt_endTime-baseline_bt_startTime, baseline_park_endTime-baseline_park_startTime],
            'Task Steps': [baseline_map_step, baseline_mcs_step, baseline_climate_step, baseline_bt_step, baseline_park_step],
            'Average HR': [baseline_map_hr_mean, baseline_mcs_hr_mean, baseline_climate_hr_mean, baseline_bt_hr_mean, baseline_park_hr_mean],
            'Average Velocity': [baseline_map_velocity_mean, baseline_mcs_velocity_mean, baseline_climate_velocity_mean, baseline_bt_velocity_mean, baseline_park_velocity_mean],
            'Average Acceleration': [baseline_map_acc_mean, baseline_mcs_acc_mean, baseline_climate_acc_mean, baseline_bt_acc_mean, baseline_park_acc_mean],
            'Average BrakeStatus': [baseline_map_bs_mean, baseline_mcs_bs_mean, baseline_climate_bs_mean, baseline_bt_bs_mean, baseline_park_bs_mean],
            'Average LaneOffset': [baseline_map_laneoff_mean, baseline_mcs_laneoff_mean, baseline_climate_laneoff_mean, baseline_bt_laneoff_mean, baseline_park_laneoff_mean],
            'Average SteeringWheelAngle': [baseline_map_swa_mean, baseline_mcs_swa_mean, baseline_climate_swa_mean, baseline_bt_swa_mean, baseline_park_swa_mean],
            'ADAS#': [baseline_map_adasNum, baseline_mcs_adasNum, baseline_climate_adasNum, baseline_bt_adasNum, baseline_park_adasNum]
        })
        update.to_excel(writer, 'Baseline Task',
            columns=['Task Name','Task Start Time', 'Task End Time', 'Task Duration', 'Task Steps', 'Average HR', 'Average Velocity', 'Average Acceleration', 'Average BrakeStatus', 'Average LaneOffset', 'Average SteeringWheelAngle', 'ADAS#'],
            index=False)
        writer.save()
